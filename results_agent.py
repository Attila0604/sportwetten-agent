"""
Ergebnis-Agent — prüft täglich ob Wetten gewonnen oder verloren haben
Läuft täglich um 10:00 Uhr und aktualisiert die Excel-Datei automatisch

REWRITE 04.05.2026:
  - FIX 5: _aehnlich war zu lasch (z.B. "Bayer" matchte "Bayern", "Wien" matchte "Wien")
           → strikteres Matching via difflib + Alias-Tabelle
  - FIX 6: Score-Positions-Fallback entfernt (Reihenfolge laut Odds API NICHT garantiert!)
  - FIX 7: Debug-Warnungen entrümpelt (max. 1 pro Bet statt pro API-Treffer)
"""
import os
import difflib
import httpx
from datetime import datetime, date, timedelta

ODDS_API_KEY = os.getenv("ODDS_API_KEY")
BASE_URL = "https://api.the-odds-api.com/v4"

SOCCER_SPORTS = [
    "soccer_germany_bundesliga",
    "soccer_england_league1",
    "soccer_epl",
    "soccer_spain_la_liga",
    "soccer_italy_serie_a",
    "soccer_france_ligue_one",
    "soccer_uefa_champs_league",
    "soccer_austria_bundesliga",
]

# Alias-Tabelle: Teams, die in Wetten und API unterschiedlich heißen.
# Schreibweise EGAL — wird beim Lookup normalisiert.
# Bei weiteren Mismatches einfach hier ergänzen.
TEAM_ALIASES = [
    # --- Österreich Bundesliga ---
    {"RB Salzburg", "FC Red Bull Salzburg", "Red Bull Salzburg"},
    {"Rapid Wien", "SK Rapid Wien", "SK Rapid"},
    {"Austria Wien", "FK Austria Wien", "FK Austria"},
    {"Sturm Graz", "SK Sturm Graz"},
    {"WSG Tirol", "WSG Swarovski Tirol"},
    {"LASK", "LASK Linz", "Linzer ASK"},
    {"Austria Klagenfurt", "SK Austria Klagenfurt"},
    {"Blau-Weiß Linz", "FC Blau-Weiß Linz"},
    {"TSV Hartberg", "Hartberg"},
    {"Wolfsberger AC", "WAC", "Wolfsberg"},
    {"Rheindorf Altach", "SCR Altach", "Altach"},
    # --- Deutschland Bundesliga (häufige Stolperfallen) ---
    {"Bayer Leverkusen", "Bayer 04 Leverkusen"},
    {"Bayern München", "FC Bayern München", "Bayern Munich"},
    {"Borussia Dortmund", "BVB Dortmund", "BV Borussia 09 Dortmund"},
    {"Borussia Mönchengladbach", "Borussia M'gladbach", "Bor. Mönchengladbach"},
    {"1. FC Köln", "FC Köln"},
    {"1. FC Kaiserslautern", "FCK Kaiserslautern"},
    # --- Spanien (häufige "Real"/"Atletico" Verwechslungen) ---
    {"Real Madrid", "Real Madrid CF"},
    {"Real Sociedad", "Real Sociedad de Fútbol"},
    {"Atlético Madrid", "Atletico Madrid", "Club Atlético de Madrid"},
    {"Athletic Bilbao", "Athletic Club"},
    # --- England (häufige "Manchester"/"United" Verwechslungen) ---
    {"Manchester United", "Man United", "Man Utd"},
    {"Manchester City", "Man City"},
    {"Newcastle United", "Newcastle"},
    {"West Ham United", "West Ham"},
    {"Tottenham Hotspur", "Tottenham", "Spurs"},
]


def _normalisiere(name: str) -> str:
    """Normalisiert Teamnamen: lowercase, Umlaute, Sonderzeichen, Präfixe entfernen"""
    if not name:
        return ""
    name = name.lower().strip()
    # Umlaute & ß
    name = name.replace("ä", "a").replace("ö", "o").replace("ü", "u")
    name = name.replace("ae", "a").replace("oe", "o").replace("ue", "u")
    name = name.replace("ß", "ss")
    # Apostrophe und Akzente vereinfachen
    name = name.replace("é", "e").replace("è", "e").replace("á", "a").replace("í", "i")
    # Häufige Präfixe entfernen (längere zuerst, nur einer wird abgeschnitten)
    for prefix in [
        "1. fc ", "1.fc ", "1 fc ",
        "fc ", "sc ", "sv ", "vfl ", "vfb ", "bv ", "tsv ", "fsv ",
        "sk ", "fk ", "wsg ", "rb ", "ask ", "atv ", "as ", "ac ",
    ]:
        if name.startswith(prefix):
            name = name[len(prefix):]
            break
    # Häufige Suffixe entfernen
    for suffix in [" fc", " sc", " sv", " cf", " ac"]:
        if name.endswith(suffix):
            name = name[: -len(suffix)]
    # Restliche Sonderzeichen
    name = name.replace(".", "").replace("-", " ").replace("'", "").replace("/", " ")
    # Mehrfache Whitespaces zusammenführen
    name = " ".join(name.split())
    return name


# Aliase einmalig vornormalisieren für schnellen Lookup
_NORMALIZED_ALIAS_GROUPS = [
    {_normalisiere(n) for n in gruppe} for gruppe in TEAM_ALIASES
]


def _alias_match(na: str, nb: str) -> bool:
    """Prüft, ob zwei normalisierte Namen über die Alias-Tabelle gematcht werden."""
    for gruppe in _NORMALIZED_ALIAS_GROUPS:
        if na in gruppe and nb in gruppe:
            return True
    return False


def _aehnlich(a: str, b: str, schwelle: float = 0.85) -> bool:
    """
    Prüft ob zwei Teamnamen das gleiche Team bezeichnen.
    
    DEUTLICH STRENGER als die alte Version:
    - Kein 4-Zeichen-Wort-Matching mehr ("Bayer"↔"Bayern", "Wien"↔"Wien" → früher TRUE!)
    - Substring nur bei mind. 70% Längen-Verhältnis (verhindert "RB"↔"Rapid Bilbao")
    - Zusätzlich Alias-Tabelle für bekannte Schreibvariationen
    """
    na, nb = _normalisiere(a), _normalisiere(b)
    if not na or not nb:
        return False
    # 1. Exakt gleich nach Normalisierung
    if na == nb:
        return True
    # 2. Bekannter Alias
    if _alias_match(na, nb):
        return True
    # 3. Substring NUR bei beidseitig ≥ 6 Zeichen UND ≥ 70% Längen-Verhältnis
    if len(na) >= 6 and len(nb) >= 6:
        kurz, lang = (na, nb) if len(na) <= len(nb) else (nb, na)
        if kurz in lang and len(kurz) / len(lang) >= 0.70:
            return True
    # 4. Difflib-Fuzzy-Match (fängt Tippfehler & kleine Schreibvarianten ab)
    if difflib.SequenceMatcher(None, na, nb).ratio() >= schwelle:
        return True
    return False


async def hole_ergebnisse(sport: str) -> list:
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(
                f"{BASE_URL}/sports/{sport}/scores/",
                params={
                    "apiKey": ODDS_API_KEY,
                    "daysFrom": 3  # FIX Bug 1: war 2, zu knapp für Spiele von vor 2+ Tagen
                }
            )
            if resp.status_code == 200:
                return resp.json()
            else:
                print(f"  API Fehler ({sport}): {resp.status_code}")
    except Exception as e:
        print(f"  Ergebnis-Fehler ({sport}): {e}")
    return []


async def hole_alle_ergebnisse() -> list:
    alle = []
    for sport in SOCCER_SPORTS:
        ergebnisse = await hole_ergebnisse(sport)
        for e in ergebnisse:
            e["liga_key"] = sport
        alle.extend(ergebnisse)
    print(f"  → Gesamt {len(alle)} Spiele von API geladen")
    return alle


def finde_ergebnis(heim: str, gast: str, ergebnisse: list) -> dict:
    """Sucht das Ergebnis für ein bestimmtes Spiel"""
    bester_kandidat = None  # Für eine einzelne Debug-Zeile, falls nichts matcht

    for spiel in ergebnisse:
        spiel_heim = spiel.get("home_team", "")
        spiel_gast = spiel.get("away_team", "")

        heim_match = _aehnlich(heim, spiel_heim)
        gast_match = _aehnlich(gast, spiel_gast)

        # Bester Heim-Treffer ohne Gast-Match merken (max. einer pro Bet)
        if heim_match and not gast_match and bester_kandidat is None:
            bester_kandidat = (spiel_heim, spiel_gast)

        if not (heim_match and gast_match):
            continue
        if not spiel.get("completed"):
            continue

        scores = spiel.get("scores") or []
        if len(scores) < 2:
            continue

        # FIX Bug 6: Score-Parsing STRIKT über Namen — kein Positions-Fallback mehr.
        # Die Odds API garantiert die Reihenfolge der scores nicht!
        heim_score = None
        gast_score = None
        for s in scores:
            name = s.get("name", "")
            if heim_score is None and _aehnlich(name, spiel_heim):
                heim_score = s.get("score")
            elif gast_score is None and _aehnlich(name, spiel_gast):
                gast_score = s.get("score")

        if heim_score is None or gast_score is None:
            print(f"  ⚠ Score-Parsing fehlgeschlagen: {spiel_heim} vs {spiel_gast}")
            continue

        try:
            heim_score = int(heim_score)
            gast_score = int(gast_score)
        except (ValueError, TypeError):
            print(f"  ⚠ Score nicht numerisch: {spiel_heim} vs {spiel_gast}")
            continue

        return {
            "heim": spiel_heim,
            "gast": spiel_gast,
            "heim_score": heim_score,
            "gast_score": gast_score,
            "endstand": f"{heim_score}:{gast_score}",
            "abgeschlossen": True,
        }

    # FIX Bug 7: Nur EINE Debug-Zeile pro Bet, statt eine pro API-Game
    if bester_kandidat:
        sh, sg = bester_kandidat
        print(f"  ⚡ Kein Match für '{heim} vs {gast}' (ähnlich gefunden: '{sh} vs {sg}')")

    return {}


def berechne_wett_ergebnis(empfehlung: str, heim_score: int, gast_score: int) -> str:
    if empfehlung == "Heimsieg":
        return "Gewonnen" if heim_score > gast_score else "Verloren"
    elif empfehlung == "Gastsieg":
        return "Gewonnen" if gast_score > heim_score else "Verloren"
    elif empfehlung == "Unentschieden":
        return "Gewonnen" if heim_score == gast_score else "Verloren"
    return "Offen"


def berechne_gewinn_verlust(status: str, einsatz: float, quote: float) -> float:
    if status == "Gewonnen":
        return round(einsatz * quote - einsatz, 2)
    elif status == "Verloren":
        return round(-einsatz, 2)
    return 0.0


def aktualisiere_kapital_sheet(ws_kapital, heute: date, tages_gv: float,
                                kapital_vorher: float, anzahl_wetten: int):
    """Schreibt täglichen Kapitalstand ins Kapital-Sheet"""
    naechste_zeile = 3
    for row in range(3, ws_kapital.max_row + 2):
        if ws_kapital.cell(row=row, column=1).value is None:
            naechste_zeile = row
            break

    kapital_neu = round(kapital_vorher + tages_gv, 2)
    roi = round((tages_gv / kapital_vorher * 100), 2) if kapital_vorher > 0 else 0.0

    ws_kapital.cell(row=naechste_zeile, column=1).value = heute.strftime("%d.%m.%Y")
    ws_kapital.cell(row=naechste_zeile, column=2).value = kapital_neu
    ws_kapital.cell(row=naechste_zeile, column=3).value = round(tages_gv, 2)
    ws_kapital.cell(row=naechste_zeile, column=4).value = anzahl_wetten
    ws_kapital.cell(row=naechste_zeile, column=5).value = roi
    ws_kapital.cell(row=naechste_zeile, column=6).value = "Auto-Update"

    print(f"  → Kapital-Sheet: {kapital_vorher:.2f}€ + {tages_gv:+.2f}€ = {kapital_neu:.2f}€")
    return kapital_neu


async def ergebnisse_aktualisieren() -> dict:
    from excel_agent import EXCEL_PATH
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    GRUEN = "FF2ECC71"
    ROT = "FFE74C3C"
    WEISS = "FFFFFFFF"

    print(f"  → Hole Spielergebnisse...")
    alle_ergebnisse = await hole_alle_ergebnisse()
    print(f"  → {len(alle_ergebnisse)} Spiele gefunden")

    if not os.path.exists(EXCEL_PATH):
        print("  ⚠ Excel-Datei nicht gefunden")
        return {"aktualisiert": 0, "gewonnen": 0, "verloren": 0}

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Meine Wetten"]
    ws_kapital = wb["Kapital"]

    aktualisiert = 0
    gewonnen = 0
    verloren = 0
    tages_gv = 0.0
    kapital_aktuell = 1000.0  # Startkapital-Fallback

    # Aktuelles Kapital aus letztem Kapital-Eintrag lesen
    for row in range(ws_kapital.max_row, 2, -1):
        val = ws_kapital.cell(row=row, column=2).value
        if val is not None:
            try:
                kapital_aktuell = float(val)
            except (ValueError, TypeError):
                pass
            break

    for row_idx in range(3, ws.max_row + 1):
        status_raw = ws.cell(row=row_idx, column=15).value

        # FIX Bug 4: Whitespace-sicherer Vergleich
        status = str(status_raw or "").strip()
        if status != "Offen":
            continue

        heim = ws.cell(row=row_idx, column=3).value
        gast = ws.cell(row=row_idx, column=4).value
        empfehlung = ws.cell(row=row_idx, column=5).value
        quote = ws.cell(row=row_idx, column=6).value or 0
        einsatz = ws.cell(row=row_idx, column=10).value or 0

        if not heim or not gast:
            continue

        ergebnis = finde_ergebnis(str(heim), str(gast), alle_ergebnisse)

        if not ergebnis.get("abgeschlossen"):
            continue

        heim_score = ergebnis["heim_score"]
        gast_score = ergebnis["gast_score"]
        endstand = ergebnis["endstand"]

        wett_status = berechne_wett_ergebnis(str(empfehlung), heim_score, gast_score)
        gv = berechne_gewinn_verlust(wett_status, float(einsatz), float(quote))
        tages_gv += gv

        # Spalte 12: Endstand (z.B. "2:1")
        ws.cell(row=row_idx, column=12).value = endstand

        # Spalte 13: Vollständiger Spielstand
        ws.cell(row=row_idx, column=13).value = f"{heim} {heim_score}:{gast_score} {gast}"

        # Spalte 14: G/V EUR
        gv_z = ws.cell(row=row_idx, column=14)
        gv_z.value = gv
        gv_z.number_format = '#,##0.00'
        gv_z.font = Font(name="Arial", bold=True, size=10,
                         color=GRUEN if gv > 0 else ROT)

        # Spalte 15: Status
        sz = ws.cell(row=row_idx, column=15)
        sz.value = wett_status
        if wett_status == "Gewonnen":
            sz.fill = PatternFill("solid", start_color=GRUEN)
            sz.font = Font(name="Arial", bold=True, color=WEISS, size=10)
            gewonnen += 1
        else:
            sz.fill = PatternFill("solid", start_color=ROT)
            sz.font = Font(name="Arial", bold=True, color=WEISS, size=10)
            verloren += 1

        aktualisiert += 1
        print(f"  ✓ {heim} vs {gast}: {endstand} → {wett_status} ({gv:+.2f}€)")

    # FIX Bug 3: Kapital-Sheet befüllen wenn Wetten aktualisiert wurden
    if aktualisiert > 0:
        aktualisiere_kapital_sheet(
            ws_kapital=ws_kapital,
            heute=date.today(),
            tages_gv=tages_gv,
            kapital_vorher=kapital_aktuell,
            anzahl_wetten=aktualisiert
        )

    wb.save(EXCEL_PATH)
    print(f"  → Excel gespeichert: {aktualisiert} Wetten aktualisiert")

    return {
        "aktualisiert": aktualisiert,
        "gewonnen": gewonnen,
        "verloren": verloren,
        "tages_gv": round(tages_gv, 2),
        "gesamt_ergebnisse": len(alle_ergebnisse),
    }
