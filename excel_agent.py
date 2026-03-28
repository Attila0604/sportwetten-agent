import os
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_PATH = os.getenv("EXCEL_PATH", "data/wett_tracker.xlsx")
STARTKAPITAL = float(os.getenv("STARTKAPITAL", "1000"))

DUNKEL    = "FF0D1117"
BLAU      = "FF1A8FE3"
GRUEN     = "FF2ECC71"
ROT       = "FFE74C3C"
GOLD      = "FFF1C40F"
WEISS     = "FFFFFFFF"
GRAU      = "FF8B949E"
HELLGRAU  = "FFF0F2F5"
NAVY      = "FF1A252F"
DUNKELGRAU = "FF2C3E50"
SCHWARZ   = "FF000000"

def rand(c="FFDDDDDD"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def kh(z, txt, fg=WEISS, bg=DUNKEL, sz=11, fett=True):
    z.value = txt
    z.font = Font(name="Arial", bold=fett, color=fg, size=sz)
    z.fill = PatternFill("solid", start_color=bg)
    z.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    z.border = rand()

def dz(z, v, fmt=None, fett=False, fg=SCHWARZ, bg=WEISS, mitte=True):
    z.value = v
    z.font = Font(name="Arial", bold=fett, color=fg, size=10)
    z.fill = PatternFill("solid", start_color=bg)
    if fmt:
        z.number_format = fmt
    z.alignment = Alignment(horizontal="center" if mitte else "left", vertical="center")
    z.border = rand()

def erstelle_excel():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Empfehlungen"
    _sheet_empfehlungen(ws1)
    ws2 = wb.create_sheet("Meine Wetten")
    _sheet_wetten(ws2)
    ws3 = wb.create_sheet("Kapital")
    _sheet_kapital(ws3)
    ws4 = wb.create_sheet("Wochenanalyse")
    _sheet_analyse(ws4, "WOCHENANALYSE")
    ws5 = wb.create_sheet("Monatsanalyse")
    _sheet_analyse(ws5, "MONATSANALYSE")
    ws6 = wb.create_sheet("Statistik")
    _sheet_statistik(ws6)
    os.makedirs(os.path.dirname(EXCEL_PATH) if os.path.dirname(EXCEL_PATH) else ".", exist_ok=True)
    wb.save(EXCEL_PATH)
    return wb

def _sheet_empfehlungen(ws):
    ws.freeze_panes = "A3"
    ws.merge_cells("A1:N1")
    t = ws["A1"]
    t.value = "BET365 FUSSBALL VALUE BETS — KI TAGESANALYSE (21:00 Uhr)"
    t.font = Font(name="Arial", bold=True, color=WEISS, size=14)
    t.fill = PatternFill("solid", start_color=DUNKEL)
    t.alignment = Alignment(horizontal="center", vertical="center")
    cols = [
        ("A","Datum",12),("B","Liga",20),("C","Heim",22),("D","Gast",22),
        ("E","Uhrzeit",10),("F","Empfehlung",15),("G","Quote",9),
        ("H","Impl.%",10),("I","Echt%",10),("J","EV%",10),
        ("K","Konfidenz",12),("L","Risiko",10),("M","Top 3?",10),("N","Begruendung",35),
    ]
    for col, hdr, w in cols:
        kh(ws[f"{col}2"], hdr)
        ws.column_dimensions[col].width = w

def _sheet_wetten(ws):
    ws.freeze_panes = "A3"
    ws.merge_cells("A1:P1")
    t = ws["A1"]
    t.value = f"MEINE WETTEN — TOP 3 AUTO-TRACKING | Startkapital: {STARTKAPITAL:.0f}EUR"
    t.font = Font(name="Arial", bold=True, color=WEISS, size=13)
    t.fill = PatternFill("solid", start_color=DUNKEL)
    t.alignment = Alignment(horizontal="center", vertical="center")
    cols = [
        ("A","Datum",12),("B","Liga",18),("C","Heim",20),("D","Gast",20),
        ("E","Empfehlung",14),("F","Quote",9),("G","Kapital EUR",12),
        ("H","Max 3%",11),("I","Kelly EUR",11),("J","Einsatz EUR",12),
        ("K","Pot.Gewinn",14),("L","Endstand",12),("M","Ergebnis",15),
        ("N","G/V EUR",14),("O","Status",12),("P","EV%",9),
    ]
    for col, hdr, w in cols:
        kh(ws[f"{col}2"], hdr)
        ws.column_dimensions[col].width = w

def _sheet_kapital(ws):
    ws.freeze_panes = "A3"
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"KAPITALVERLAUF — Startkapital: {STARTKAPITAL:.0f}EUR"
    t.font = Font(name="Arial", bold=True, color=WEISS, size=14)
    t.fill = PatternFill("solid", start_color=DUNKEL)
    t.alignment = Alignment(horizontal="center", vertical="center")
    for col, hdr, w in [("A","Datum",14),("B","Kapital",16),("C","Tagesgewinn",16),("D","Wetten",14),("E","ROI%",14),("F","Notiz",25)]:
        kh(ws[f"{col}2"], hdr)
        ws.column_dimensions[col].width = w

def _sheet_analyse(ws, titel):
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = f"{titel} — KI AUSWERTUNG"
    t.font = Font(name="Arial", bold=True, color=WEISS, size=14)
    t.fill = PatternFill("solid", start_color=DUNKEL)
    t.alignment = Alignment(horizontal="center", vertical="center")
    for col, hdr, w in [("A","Zeitraum",20),("B","KI-Analyse",80),("C","Wetten",10),("D","ROI%",12)]:
        kh(ws[f"{col}2"], hdr)
        ws.column_dimensions[col].width = w

def _sheet_statistik(ws):
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "GESAMTSTATISTIK"
    t.font = Font(name="Arial", bold=True, color=WEISS, size=16)
    t.fill = PatternFill("solid", start_color=DUNKEL)
    t.alignment = Alignment(horizontal="center", vertical="center")
    labels = [
        (3,"Startkapital EUR"),(4,"Aktuelles Kapital EUR"),
        (5,"Gesamt G/V EUR"),(6,"ROI %"),(7,"Gesamt Wetten"),
        (8,"Gewonnen"),(9,"Verloren"),(10,"Offen"),(11,"Trefferquote %"),
        (12,"Durchschn. Quote"),(13,"Durchschn. Einsatz"),(14,"Beste Wette"),(15,"Schlechteste Wette"),
    ]
    formeln = {
        # FIX Bug 3: STARTKAPITAL als float speichern, nicht als str → verhindert #WERT!
        3: STARTKAPITAL,
        4: "=C3+SUM('Meine Wetten'!N3:N10000)",
        5: "=C4-C3", 6: "=IF(C3>0,(C4-C3)/C3*100,0)",
        7: "=COUNTA('Meine Wetten'!A3:A10000)",
        8: "=COUNTIF('Meine Wetten'!O3:O10000,\"Gewonnen\")",
        9: "=COUNTIF('Meine Wetten'!O3:O10000,\"Verloren\")",
        10: "=COUNTIF('Meine Wetten'!O3:O10000,\"Offen\")",
        11: "=IF(C7>0,C8/C7*100,0)",
        12: "=IF(C7>0,AVERAGEIF('Meine Wetten'!F3:F10000,\">0\"),0)",
        13: "=IF(C7>0,AVERAGE('Meine Wetten'!J3:J10000),0)",
        14: "=IF(C7>0,MAX('Meine Wetten'!N3:N10000),0)",
        15: "=IF(C7>0,MIN('Meine Wetten'!N3:N10000),0)",
    }
    for row, lbl in labels:
        l = ws.cell(row=row, column=1, value=lbl)
        l.font = Font(name="Arial", bold=True, color=WEISS, size=11)
        l.fill = PatternFill("solid", start_color=DUNKELGRAU)
        l.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        l.border = rand("FF555555")
        ws.column_dimensions["A"].width = 28
        v = ws.cell(row=row, column=3)
        v.value = formeln.get(row, "")
        v.font = Font(name="Arial", bold=True, size=13, color=BLAU)
        v.fill = PatternFill("solid", start_color=HELLGRAU)
        v.alignment = Alignment(horizontal="center", vertical="center")
        v.border = rand()
        ws.column_dimensions["C"].width = 22


# ─── DUPLIKAT-HELPER ──────────────────────────────────────────────────────────

def _empfehlung_existiert(ws, datum: str, heim: str, gast: str) -> bool:
    """Prüft ob eine Empfehlung für Datum+Heim+Gast bereits existiert"""
    for row in ws.iter_rows(min_row=3, values_only=True):
        if (str(row[0] or "").strip() == datum and
                str(row[2] or "").strip().lower() == heim.strip().lower() and
                str(row[3] or "").strip().lower() == gast.strip().lower()):
            return True
    return False


def _wette_existiert(ws, datum: str, heim: str, gast: str) -> bool:
    """Prüft ob eine Wette für Datum+Heim+Gast bereits existiert"""
    for row in ws.iter_rows(min_row=3, values_only=True):
        if (str(row[0] or "").strip() == datum and
                str(row[2] or "").strip().lower() == heim.strip().lower() and
                str(row[3] or "").strip().lower() == gast.strip().lower()):
            return True
    return False


# ─── HAUPT-FUNKTIONEN ─────────────────────────────────────────────────────────

def empfehlungen_hinzufuegen(result: dict, datum: str = None):
    if not os.path.exists(EXCEL_PATH):
        erstelle_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Empfehlungen"]
    heute = datum or date.today().strftime("%d.%m.%Y")
    alle = result.get("alle_empfehlungen", [])
    top3_ids = {e.get("spiel_id") for e in result.get("top3", [])}
    zeile = max(ws.max_row + 1, 3)
    eingefuegt = 0
    uebersprungen = 0

    for i, emp in enumerate(alle):
        heim = emp.get("heim", "")
        gast = emp.get("gast", "")

        # FIX Bug 1: Duplikat-Check — gleiche Empfehlung nicht doppelt eintragen
        if _empfehlung_existiert(ws, heute, heim, gast):
            print(f"  ⏭ Duplikat übersprungen: {heim} vs {gast}")
            uebersprungen += 1
            continue

        bg = WEISS if eingefuegt % 2 == 0 else HELLGRAU
        r = zeile + eingefuegt
        ist_top3 = emp.get("spiel_id") in top3_ids
        ev_p = round(emp.get("ev", 0) * 100, 2)
        row_bg = "FFD4EDDA" if ist_top3 else bg
        vals = [
            (1,heute,None,False,SCHWARZ),(2,emp.get("liga",""),None,False,SCHWARZ),
            (3,heim,None,True,NAVY),(4,gast,None,True,NAVY),
            (5,emp.get("zeit","")[-5:] if len(emp.get("zeit",""))>5 else "",None,False,SCHWARZ),
            (6,emp.get("empfehlung",""),None,True,SCHWARZ),
            (7,emp.get("quote",0),'#,##0.00',True,BLAU),
            (8,round(emp.get("implizite_wahrscheinlichkeit",0)*100,1),'#,##0.0',False,SCHWARZ),
            (9,round(emp.get("echte_wahrscheinlichkeit",0)*100,1),'#,##0.0',False,SCHWARZ),
            (10,ev_p,'#,##0.00',True,GRUEN if ev_p>0 else ROT),
            (11,emp.get("konfidenz",0),None,True,SCHWARZ),
            (12,emp.get("risiko",""),None,False,SCHWARZ),
            (13,"TOP 3" if ist_top3 else "",None,True,GOLD if ist_top3 else SCHWARZ),
            (14,emp.get("begruendung",""),None,False,SCHWARZ),
        ]
        for col,val,fmt,fett,fg in vals:
            dz(ws.cell(row=r,column=col),val,fmt,fett,fg,row_bg)
        ws.row_dimensions[r].height = 20
        eingefuegt += 1

    wb.save(EXCEL_PATH)
    print(f"  → Empfehlungen: {eingefuegt} neu, {uebersprungen} Duplikate übersprungen")


def top3_als_wetten_eintragen(top3: list, aktuelles_kapital: float, datum: str = None):
    if not os.path.exists(EXCEL_PATH):
        erstelle_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Meine Wetten"]
    heute = datum or date.today().strftime("%d.%m.%Y")
    zeile = max(ws.max_row + 1, 3)
    eingefuegt = 0
    uebersprungen = 0

    for i, emp in enumerate(top3):
        heim = emp.get("heim", "")
        gast = emp.get("gast", "")

        # FIX Bug 2: Duplikat-Check — gleiche Wette nicht doppelt eintragen
        if _wette_existiert(ws, heute, heim, gast):
            print(f"  ⏭ Wett-Duplikat übersprungen: {heim} vs {gast}")
            uebersprungen += 1
            continue

        bg = WEISS if eingefuegt % 2 == 0 else HELLGRAU
        r = zeile + eingefuegt
        einsatz = emp.get("empfohlener_einsatz", 0)
        quote = emp.get("quote", 0)
        pot_gewinn = round(einsatz * quote - einsatz, 2)
        ev_p = round(emp.get("ev", 0) * 100, 2)
        vals = [
            (1,heute,None,False,SCHWARZ),(2,emp.get("liga",""),None,False,SCHWARZ),
            (3,heim,None,True,NAVY),(4,gast,None,True,NAVY),
            (5,emp.get("empfehlung",""),None,True,SCHWARZ),
            (6,quote,'#,##0.00',True,BLAU),
            (7,round(aktuelles_kapital,2),'#,##0.00',False,SCHWARZ),
            (8,round(emp.get("max_einsatz",0),2),'#,##0.00',False,GRAU),
            (9,round(emp.get("kelly_einsatz",0),2),'#,##0.00',False,BLAU),
            (10,einsatz,'#,##0.00',True,NAVY),
            (11,pot_gewinn,'#,##0.00',True,GRUEN),
            (12,"- : -",None,False,SCHWARZ),(13,"Ausstehend",None,False,SCHWARZ),
            (14,0.0,'#,##0.00',True,SCHWARZ),(15,"Offen",None,True,SCHWARZ),
            (16,ev_p,'#,##0.00',False,GRUEN if ev_p>0 else ROT),
        ]
        for col,val,fmt,fett,fg in vals:
            dz(ws.cell(row=r,column=col),val,fmt,fett,fg,bg)
        ws.row_dimensions[r].height = 20
        eingefuegt += 1

    wb.save(EXCEL_PATH)
    print(f"  → Wetten: {eingefuegt} neu, {uebersprungen} Duplikate übersprungen")


def kapital_eintragen(kapital: float, tagesgewinn: float, anzahl_wetten: int, notiz: str = ""):
    if not os.path.exists(EXCEL_PATH):
        erstelle_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Kapital"]
    zeile = max(ws.max_row + 1, 3)
    roi = round(tagesgewinn / (kapital - tagesgewinn) * 100, 2) if (kapital - tagesgewinn) > 0 else 0
    vals = [date.today().strftime("%d.%m.%Y"), round(kapital,2), round(tagesgewinn,2), anzahl_wetten, roi, notiz]
    fmts = [None,'#,##0.00','#,##0.00',None,'#,##0.00',None]
    for col,(val,fmt) in enumerate(zip(vals,fmts),1):
        fg = GRUEN if col==3 and tagesgewinn>0 else ROT if col==3 and tagesgewinn<0 else SCHWARZ
        dz(ws.cell(row=zeile,column=col),val,fmt,col in [2,3],fg)
    ws.row_dimensions[zeile].height = 20
    wb.save(EXCEL_PATH)

def analyse_eintragen(zeitraum: str, analyse_text: str, anzahl_wetten: int, roi: float, monatlich: bool = False):
    if not os.path.exists(EXCEL_PATH):
        erstelle_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Monatsanalyse" if monatlich else "Wochenanalyse"]
    zeile = max(ws.max_row + 1, 3)
    ws.row_dimensions[zeile].height = 80
    dz(ws.cell(row=zeile,column=1),zeitraum,None,True)
    z = ws.cell(row=zeile,column=2)
    z.value = analyse_text
    z.font = Font(name="Arial",size=10)
    z.alignment = Alignment(vertical="top",wrap_text=True)
    z.border = rand()
    ws.column_dimensions["B"].width = 80
    dz(ws.cell(row=zeile,column=3),anzahl_wetten,None,True)
    dz(ws.cell(row=zeile,column=4),roi,'#,##0.00',True,GRUEN if roi>0 else ROT)
    wb.save(EXCEL_PATH)

def wette_aktualisieren(zeile_idx: int, endstand: str, ergebnis: str, gewinn_verlust: float, status: str):
    if not os.path.exists(EXCEL_PATH):
        return
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Meine Wetten"]
    r = zeile_idx + 2
    ws.cell(row=r,column=12).value = endstand
    ws.cell(row=r,column=13).value = ergebnis
    gv_z = ws.cell(row=r,column=14)
    gv_z.value = round(gewinn_verlust,2)
    gv_z.font = Font(name="Arial",bold=True,size=10,color=GRUEN if gewinn_verlust>0 else ROT)
    gv_z.number_format = '#,##0.00'
    sz = ws.cell(row=r,column=15)
    sz.value = status
    if status == "Gewonnen":
        sz.fill = PatternFill("solid",start_color=GRUEN)
        sz.font = Font(name="Arial",bold=True,color=WEISS,size=10)
    elif status == "Verloren":
        sz.fill = PatternFill("solid",start_color=ROT)
        sz.font = Font(name="Arial",bold=True,color=WEISS,size=10)
    wb.save(EXCEL_PATH)

def get_statistik() -> dict:
    if not os.path.exists(EXCEL_PATH):
        erstelle_excel()
        return _leer()
    try:
        wb = load_workbook(EXCEL_PATH,data_only=True)
        ws = wb["Meine Wetten"]
        wetten = []
        for row in ws.iter_rows(min_row=3,values_only=True):
            if row[0]:
                wetten.append({
                    "datum":str(row[0]),"liga":row[1],"heim":row[2],"gast":row[3],
                    "empfehlung":row[4],"quote":row[5] or 0,"kapital":row[6] or 0,
                    "einsatz":row[9] or 0,"pot_gewinn":row[10] or 0,"endstand":row[11],
                    "ergebnis":row[12],"gewinn_verlust":row[13] or 0,
                    "status":row[14] or "Offen","ev":row[15] or 0,
                })
        gesamt = len(wetten)
        gew = sum(1 for w in wetten if w["status"]=="Gewonnen")
        verl = sum(1 for w in wetten if w["status"]=="Verloren")
        offen = sum(1 for w in wetten if w["status"]=="Offen")
        ges_einsatz = sum(w["einsatz"] for w in wetten)
        ges_gv = sum(w["gewinn_verlust"] for w in wetten)
        return {
            "gesamt_wetten":gesamt,"gewonnen":gew,"verloren":verl,"offen":offen,
            "trefferquote":round(gew/gesamt*100,1) if gesamt>0 else 0,
            "gesamt_einsatz":round(ges_einsatz,2),"gesamt_gv":round(ges_gv,2),
            "roi":round(ges_gv/STARTKAPITAL*100,2) if STARTKAPITAL>0 else 0,
            "startkapital":STARTKAPITAL,"aktuelles_kapital":round(STARTKAPITAL+ges_gv,2),
            "letzte_wetten":wetten[-10:][::-1],"alle_wetten":wetten,
        }
    except Exception as e:
        print(f"Statistik-Fehler: {e}")
        return _leer()

def get_kapital() -> float:
    return get_statistik().get("aktuelles_kapital", STARTKAPITAL)

def _leer():
    return {
        "gesamt_wetten":0,"gewonnen":0,"verloren":0,"offen":0,
        "trefferquote":0,"gesamt_einsatz":0,"gesamt_gv":0,"roi":0,
        "startkapital":STARTKAPITAL,"aktuelles_kapital":STARTKAPITAL,
        "letzte_wetten":[],"alle_wetten":[],
    }
